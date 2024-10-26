import numpy as np
from random import randint, choice, choices
import copy
import math
import bisect
import json
import networkx as nx
import openpyxl
import matplotlib.pyplot as plt

def get_value_from_excel(file_path, operator, cloud):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Find the column index for the given cloud
    cloud_column_index = None
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=2):
        for cell in col:
            if cell.value == cloud:
                cloud_column_index = cell.column
                break
    
    if cloud_column_index is None:
        raise ValueError(f"Cloud '{cloud}' not found in the sheet")
    
    # Find the row index for the given operator
    operator_row_index = None
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == operator:
                operator_row_index = cell.row
                break
    
    if operator_row_index is None:
        raise ValueError(f"Operator '{operator}' not found in the sheet")
    
    # Get the value from the appropriate cell
    value = sheet.cell(row=operator_row_index, column=cloud_column_index).value
    
    # Multiply by a factor of 100 (like the DAG approach)
    return value * 100

def get_variables(tmp_variables, tmp):
    return tmp_variables.get(tmp, "Invalid tmp value")

def read_json_file(filename):
    # Initialize empty lists to store the data
    from_list = []
    to_list = []
    latency_list = []

    # Read the JSON file
    with open(filename, 'r') as file:
        data = json.load(file)

    # Iterate over the links and extract data
    for link in data['links']:
        from_list.append(link['from'])
        to_list.append(link['to'])
        latency_list.append(link['latency'])

    return from_list, to_list, latency_list


def read_first_row(file_path):
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Select the active sheet
    sheet = workbook.active
    
    # Read the first row
    first_row = sheet[1]
    
    # Store values of non-empty cells in a list
    values = [cell.value for cell in first_row if cell.value is not None]
    
    return values

def compute_shortest_path_dijkstra(G, source, target):
    try:
        path = nx.dijkstra_path(G, source=source, target=target, weight='weight')
        cost = nx.dijkstra_path_length(G, source=source, target=target, weight='weight')
        return path, cost
    except nx.NetworkXNoPath:
        return None, float('inf')

def create_com_graph(node1_lst, node2_lst, com_lat_list):
    G = nx.Graph()
    for i in range(0, len(node1_lst)):
        G.add_edge(node1_lst[i], node2_lst[i], weight = com_lat_list[i])
    return G


def calculate_force(operator_positions, k):
    forces = np.zeros_like(operator_positions, dtype=np.float64)
    for i, (op1_idx, op2_idx) in enumerate(workflow_edges_idx):
        op1_pos = operator_positions[op1_idx]
        op2_pos = operator_positions[op2_idx]
        diff = op2_pos - op1_pos
        distance = np.linalg.norm(diff)
        force = k * diff / distance  # Hooke's Law
        forces[op1_idx] += force
        forces[op2_idx] -= force
    return forces

def spring_relaxation():
    global operator_positions
    for _ in range(num_iterations):
        forces = calculate_force(operator_positions, k)
        operator_positions += forces * damping
        
        # Apply constraints for fixed operators
        for idx in fixed_operator_indices:
            operator_positions[idx] = fixed_node_position


# Define the workflow operators
operators = [
    "bloomFilter", "interpolation", "azureTableInsert", 
    "rangeFilter", "annotate", 
    "source", "join", "mqttPublish", "senmlParse", "csvToSenml", "sink"
]

# Define the number of physical nodes (network nodes)
num_nodes = 31

# Spring relaxation parameters
num_iterations = 74737
k = 1.0  # Spring constant
damping = 0.1  # Damping factor

# Define the workflow edges (workflow)
workflow_edges = [
    ("source", "senmlParse"),
    ("senmlParse", "rangeFilter"),
    ("rangeFilter", "bloomFilter"),
    ("bloomFilter", "interpolation"),
    ("interpolation", "join"),
    ("join", "annotate"),
    ("annotate", "csvToSenml"),
    ("annotate", "azureTableInsert"),
    ("csvToSenml", "mqttPublish"),
    ("azureTableInsert", "sink"),
    ("mqttPublish", "sink")
]

# Map operators to indices, e.g., bloomFilter -> 0, interpolation -> 1 etc.
operator_indices = {op: idx for idx, op in enumerate(operators)}

# Convert workflow edges to indices, e.g., (5, 8), (8, 3), ...
workflow_edges_idx = [(operator_indices[op1], operator_indices[op2]) for op1, op2 in workflow_edges]

# Define the coordinates
# Fixed positions for physical nodes in a 2D space. Use the json files provided by Banelas (there is inconsistency due to coordinates)
# The first pair is the first network node in the xlsx file.
node_values_7 = [[0, 0], [-3, 0], [-3, 2], [-3, -10], [8, 0], [8, 8], [8, -3]]
node_values_15 = [[0, 0], [-1, 0], [-3, 0], [-3, 6], [-12, 0], [-1, -4], [-2, -4], [-1, -9], [9, 0], [13, 0], [13, 7], [23, 0], [9, -1], [19, -1], [9, -9]]
node_values_31 = [[0, 0], [-9, 0], [-10, 0], [-14, 0], [-14, 4], [-17, 0], [-10, -1], [-11, -1], [-10, -3], [-9, 10], [-9, 17], [-15, 17], [-9, 20], [-2, 10], [-2, 9], [-2, 15], [8, 0], [14, 0], [19, 0], [19, 4], [24, 0], [14, 8], [14, 13], [5, 8], [8, -9], [8, -10], [8, -16], [15, -10], [2, -9], [2, -7], [2, -12]]
variables_node = {7: node_values_7, 15: node_values_15, 31: node_values_31}
node_values = get_variables(variables_node, num_nodes)
node_positions = np.array(node_values)

# Initial positions for operators in a 2D space. In detail, we assign each operator at the position that the best configuration exist, e.g., if 'source' has the lowest latency in 'cloud' then the coordinates of 'source' are equal to the coordinates of 'cloud'
# The first pair is for the operator with idx 0 etc.
op_values_7 = [[-3, 0], [-3, 0], [0, 0], [-3, 2], [-3, 2], [8, 0], [8, 0], [-3, 2], [0, 0], [0, 0], [-3, 0]]
op_values_15 = [[-1, 0], [-3, 0], [0, 0], [0, 0], [0, 0], [0, 0], [-12, 0], [13, 0], [0, 0], [0, 0], [0, 0]]
op_values_31 = [[14, 13], [-2, 10], [0, 0], [-14, 4], [-2, 9], [5, 8], [-14, 4], [5, 8], [0, 0], [0, 0], [5, 8]]
variables_op = {7: op_values_7, 15: op_values_15, 31: op_values_31}
op_values = get_variables(variables_op, num_nodes)
operator_positions = np.array(op_values, dtype=np.float64)

# Operators that must be assigned to physical node 0 (i.e., cloud)
fixed_operators = ["senmlParse", "csvToSenml", "azureTableInsert"]
fixed_operator_indices = [operator_indices[op] for op in fixed_operators]
fixed_node_position = node_positions[0]

# Run the spring relaxation algorithm
spring_relaxation()

# After the algorithm for each operator compute the closest network node and make the mapping 
best_mapping = [np.argmin(np.linalg.norm(node_positions - pos, axis=1)) for pos in operator_positions]
print("Best Mapping:", best_mapping)

# Plot the final positions
plt.figure(figsize=(10, 10))
plt.scatter(node_positions[:, 0], node_positions[:, 1], color='blue', label='Physical Nodes')
plt.scatter(operator_positions[:, 0], operator_positions[:, 1], color='red', label='Operators')
for i, txt in enumerate(operators):
    plt.annotate(txt, (operator_positions[i][0], operator_positions[i][1]))
for i, txt in enumerate(range(num_nodes)):
    plt.annotate(txt, (node_positions[i][0], node_positions[i][1]))
plt.legend()
plt.show()

# Print the resulted mapping 
file_path = "datasets/etl_" + str(num_nodes) + "_dataflow.xlsx"
non_empty_values = read_first_row(file_path)
for i in range(0, len(operators)):
    print(f"Operator: {operators[i]} \t -----> {non_empty_values[best_mapping[i]]}")

# Communication costs part
# Extract the information regarding communications latencies for this simulation from the appropriate JSON file
json_file_path = "network_sample/etl_dataflow/" + str(num_nodes) + "_1/links_" + str(num_nodes) + "_1.json"  # Provide the path to your JSON file
from_list, to_list, latency_list = read_json_file(json_file_path)

# Create the communication graph
global COM_GRAPH
COM_GRAPH = create_com_graph(from_list, to_list, latency_list)

# The file that contains the data of latencies of executing some operator at some network node
file_path = "datasets/etl_" + str(num_nodes) + "_dataflow.xlsx"

total_latency = 0

###########################################
# SUM scenario
###########################################

total_latency += get_value_from_excel(file_path, 'source', non_empty_values[best_mapping[operators.index('source')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('source')]], non_empty_values[best_mapping[operators.index('senmlParse')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'senmlParse', non_empty_values[best_mapping[operators.index('senmlParse')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('senmlParse')]], non_empty_values[best_mapping[operators.index('rangeFilter')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'rangeFilter', non_empty_values[best_mapping[operators.index('rangeFilter')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('rangeFilter')]], non_empty_values[best_mapping[operators.index('bloomFilter')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'bloomFilter', non_empty_values[best_mapping[operators.index('bloomFilter')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('bloomFilter')]], non_empty_values[best_mapping[operators.index('interpolation')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'interpolation', non_empty_values[best_mapping[operators.index('interpolation')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('interpolation')]], non_empty_values[best_mapping[operators.index('join')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'join', non_empty_values[best_mapping[operators.index('join')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('join')]], non_empty_values[best_mapping[operators.index('annotate')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'annotate', non_empty_values[best_mapping[operators.index('annotate')]])

tmp1 = total_latency
tmp2 = total_latency

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('annotate')]], non_empty_values[best_mapping[operators.index('csvToSenml')]])
tmp1 += tmp
tmp1 += get_value_from_excel(file_path, 'csvToSenml', non_empty_values[best_mapping[operators.index('csvToSenml')]])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('csvToSenml')]], non_empty_values[best_mapping[operators.index('mqttPublish')]])
tmp1 += tmp
tmp1 += get_value_from_excel(file_path, 'mqttPublish', non_empty_values[best_mapping[operators.index('mqttPublish')]])

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('annotate')]], non_empty_values[best_mapping[operators.index('azureTableInsert')]])
tmp2 += tmp
tmp2 += get_value_from_excel(file_path, 'azureTableInsert', non_empty_values[best_mapping[operators.index('azureTableInsert')]])

total_latency = tmp1 + tmp2
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('azureTableInsert')]], non_empty_values[best_mapping[operators.index('sink')]])
total_latency += tmp
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('mqttPublish')]], non_empty_values[best_mapping[operators.index('sink')]])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'sink', non_empty_values[best_mapping[operators.index('sink')]])


###########################################
# MAX scenario
###########################################

# total_latency += get_value_from_excel(file_path, 'source', non_empty_values[best_mapping[operators.index('source')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('source')]], non_empty_values[best_mapping[operators.index('senmlParse')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'senmlParse', non_empty_values[best_mapping[operators.index('senmlParse')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('senmlParse')]], non_empty_values[best_mapping[operators.index('rangeFilter')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'rangeFilter', non_empty_values[best_mapping[operators.index('rangeFilter')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('rangeFilter')]], non_empty_values[best_mapping[operators.index('bloomFilter')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'bloomFilter', non_empty_values[best_mapping[operators.index('bloomFilter')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('bloomFilter')]], non_empty_values[best_mapping[operators.index('interpolation')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'interpolation', non_empty_values[best_mapping[operators.index('interpolation')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('interpolation')]], non_empty_values[best_mapping[operators.index('join')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'join', non_empty_values[best_mapping[operators.index('join')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('join')]], non_empty_values[best_mapping[operators.index('annotate')]])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'annotate', non_empty_values[best_mapping[operators.index('annotate')]])

# tmp1 = total_latency
# tmp2 = total_latency

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('annotate')]], non_empty_values[best_mapping[operators.index('csvToSenml')]])
# tmp1 += tmp
# tmp1 += get_value_from_excel(file_path, 'csvToSenml', non_empty_values[best_mapping[operators.index('csvToSenml')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('csvToSenml')]], non_empty_values[best_mapping[operators.index('mqttPublish')]])
# tmp1 += tmp
# tmp1 += get_value_from_excel(file_path, 'mqttPublish', non_empty_values[best_mapping[operators.index('mqttPublish')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('mqttPublish')]], non_empty_values[best_mapping[operators.index('sink')]])
# tmp1 += tmp

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('annotate')]], non_empty_values[best_mapping[operators.index('azureTableInsert')]])
# tmp2 += tmp
# tmp2 += get_value_from_excel(file_path, 'azureTableInsert', non_empty_values[best_mapping[operators.index('azureTableInsert')]])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, non_empty_values[best_mapping[operators.index('azureTableInsert')]], non_empty_values[best_mapping[operators.index('sink')]])
# tmp2 += tmp

# total_latency = max(tmp1, tmp2)

# total_latency += get_value_from_excel(file_path, 'sink', non_empty_values[best_mapping[operators.index('sink')]])



# Print the latency of the resulted topology
print(f"\nThe latency of this topology is: {total_latency}")
