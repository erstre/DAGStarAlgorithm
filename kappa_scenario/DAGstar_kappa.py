from random import randint, choice, choices
import copy
import math
import json
import networkx as nx
import openpyxl
import heapq
import argparse
import pandas as pd
import sys

# Class of nodes 
class Node:
	def __init__(self, name, real, heuristic, operator, location, framework):
		self.name = name
		self.children = {}
		self.parents = []
		self.heuristic = heuristic
		self.operator_code = operator
		self.location = location
		self.framework = framework
		self.real_cost = real

	def add_child(self, child_name, weight):
		self.children[child_name] = weight

	def add_parent(self, parent_name):
		self.parents.append(parent_name)

	def get_children(self):
		return list(self.children.keys())

	def get_parents(self):
		return self.parents

	def get_operator(self):
		return self.operator_code

	def get_location(self):
		return self.location

	def get_framework(self):
		return self.framework

	def get_real_cost(self):
		return self.real_cost

	def get_heuristic_cost(self):
		return self.heuristic

	def set_location(self, location):
		self.location = location

	def set_framework(self, framework):
		self.framework = framework

	def set_real_cost(self, cost):
		self.real_cost = cost

	def set_heuristic_cost(self, cost):
		self.heuristic = cost

	def __str__(self):
		return f"(Node {self.name}) Operator {self.operator_code} in location {self.location} and framework {self.framework}, Real: {self.real_cost}, Heuristic: {self.heuristic}, Children: {self.children}, Parents: {self.parents}"

# Class of Directed (weighted) Graph
class DirectedWeightedGraph:
	def __init__(self):
		self.nodes = {}
		self.estimated_cost = -1

	def set_counterQ(self, num):
		self.counteQ = num

	def get_counterQ(self):
		return self.counteQ

	def set_estimated_cost_of_graph(self, cost):
		self.estimated_cost = cost

	def get_estimated_cost_of_graph(self):
		return self.estimated_cost

	def add_node(self, name, operator, real, heuristic, location, framework):
		if name not in self.nodes:
			self.nodes[name] = Node(name, operator, real, heuristic, location, framework)

	def add_edge(self, parent, child, weight=1):
		if parent in self.nodes and child in self.nodes:
			self.nodes[parent].add_child(child, weight)
			self.nodes[child].add_parent(parent)
		else:
			print("Node not found in the graph.")

	def delete_node(self, node_name):
		if node_name in self.nodes:
			# Delete edges connected to the node
			for parent in self.nodes[node_name].get_parents():
				del self.nodes[parent].children[node_name]

			for child in self.nodes[node_name].get_children():
				del self.nodes[child].parents[self.nodes[child].parents.index(node_name)]

			# Delete the node itself
			del self.nodes[node_name]

		else:
			print("Node not found in the graph.")

	def get_children(self, node_name):
		if node_name in self.nodes:
			return self.nodes[node_name].get_children()
		else:
			print("Node not found in the graph.")
			return []

	def get_parents(self, node_name):
		if node_name in self.nodes:
			return self.nodes[node_name].get_parents()
		else:
			print("Node not found in the graph.")
			return []

	def get_all_nodes(self):
		return list(self.nodes.keys())

	def get_root_nodes(self):
		root_nodes = []
		for node_name in self.nodes:
			if not self.nodes[node_name].get_parents():
				root_nodes.append(node_name)
		return root_nodes

	def get_leaf_nodes(self):
		leaf_nodes = []
		for node_name in self.nodes:
			if not self.nodes[node_name].get_children():
				leaf_nodes.append(node_name)
		return leaf_nodes

	def __eq__(self, other):
		# Check if they have the same nodes
		if set(self.get_all_nodes()) != set(other.get_all_nodes()):
			return False

		# Check if each node has the same parents, children, location, and framework
		for node_name in self.get_all_nodes():
			node1 = self.nodes[node_name]
			node2 = other.nodes[node_name]

			if set(node1.get_parents()) != set(node2.get_parents()):
				return False

			if set(node1.get_children()) != set(node2.get_children()):
				return False

			if node1.get_location() != node2.get_location():
				return False

			if node1.get_framework() != node2.get_framework():
				return False

		# If all checks passed, the graphs are equal
		return True

	def __hash__(self):
		tmp_str = ""
		for node_name in self.get_all_nodes():
			for ch in self.nodes[node_name].get_children():
				tmp_str += str(ch)
			for p in self.nodes[node_name].get_parents():
				tmp_str += str(p)
			tmp_str += str(self.nodes[node_name].get_location())
			tmp_str += str(self.nodes[node_name].get_framework())
			tmp_str += str(self.nodes[node_name].get_operator())
		return hash(tmp_str)

	def __str__(self):
		return "\n".join([str(self.nodes[node_name]) for node_name in self.nodes])


def extract_site_names_from_json(json_file_path):
    site_names = []

    # Open the JSON file
    with open(json_file_path, 'r') as file:
        # Load JSON data
        data = json.load(file)

        # Check if "sites" key exists in the data
        if "sites" in data:
            # Iterate over each site dictionary
            for site in data["sites"]:
                # Check if "siteName" key exists in the site dictionary
                if "siteName" in site:
                    # Append the value of "siteName" to the list
                    site_names.append(site["siteName"])
                else:
                    print("Warning: 'siteName' key not found in one of the site dictionaries.")
        else:
            print("Error: 'sites' key not found in the JSON data.")

    return site_names

def read_json_file(filename):
    # Initialize empty lists to store the data
    from_list = []
    to_list = []
    latency_list = []

    # Read the JSON file
    with open(filename, 'r') as file:
        data = json.load(file)

    # Iterate over the links and extract data
    for link in data['links']:
        from_list.append(link['from'])
        to_list.append(link['to'])
        latency_list.append(link['latency'])

    return from_list, to_list, latency_list

def read_row(excel_file, row_number):
    row_values = []

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Iterate over the specified row and store values in the list
    for col in range(2, sheet.max_column + 1):  # Iterate over all columns
        value = sheet.cell(row=row_number, column=col).value
        row_values.append(value)

    return row_values

def get_path_cost(graph, node1, node2):
    try:
        # Using Dijkstra's algorithm to find the shortest path
        shortest_path = nx.dijkstra_path(graph, node1, node2)
        
        # Calculating the cost of the path
        path_cost = sum(graph[shortest_path[i]][shortest_path[i+1]]['weight'] for i in range(len(shortest_path) - 1))
        
        return path_cost, shortest_path
    except nx.NetworkXNoPath:
        return "No path exists between the given nodes."

def create_com_graph(node1_lst, node2_lst, com_lat_list):
    G = nx.Graph()
    for i in range(0, len(node1_lst)):
        G.add_edge(node1_lst[i], node2_lst[i], weight = com_lat_list[i])
    return G

# Based on the CODE, i.e., operator this function return the minimum cost of all configurations (given a specific operator type)
def find_min_cost_of_conf(node_name, graph, operator_costs):
	operator_name = graph.nodes[node_name].get_operator()

	# Check if the operator is 'END' first
	if operator_name == 'END':
	    return 0

	# Fetch the minimum cost for the operator from operator_costs
	if operator_name in operator_costs:
	    return float(min(operator_costs[operator_name]))
	else:
	    raise ValueError(f"Unknown operator: {operator_name}")

# Function that returns the real cost based on the CODE of the operator and a specific configuration
def get_real_cost_of_conf(op_code, i, j, operator_costs):
    # Get the list of costs for the given op_code from the dictionary
    costs = operator_costs.get(op_code)
    
    if costs and i < len(costs):  # Ensure the costs exist and index is valid
        return costs[i]
    
    # Return -1 if op_code is not in the dictionary or index is invalid
    return -1

# Returns the heuristic cost of a given node that is contained in our plan of interest
# This function is called by bfs() because in order to compute the heuristic of a node, we have to compute the heuristic of its father(s) earlier 
def calculate_heuristic(node_name, end_node, start_node, graph, operator_costs):
	# In case our node of inerest is either the END_NODE or the START_NODE then its heuristic score is equal to 0
	if node_name == end_node or node_name == start_node:
		return 0
	else:
		# Get the node of interest
		my_node = graph.nodes[node_name]
		
		# Find the maximum heuristic score of its parents
		max_heuristic_of_my_parents = -1
		for my_parents in my_node.get_parents():
			val1 = graph.nodes[my_parents].heuristic
			val2 = find_min_cost_of_conf(my_parents, graph, operator_costs)
			if max_heuristic_of_my_parents <= val1 + val2:
				max_heuristic_of_my_parents = val1 + val2
		return max_heuristic_of_my_parents

# Function for computing the heuristics in the right order
def bfs(graph, end_node, start_node, operator_costs):
	print("\n=================================")
	print("Computing heuristics...")
	print("=================================")
	
	# Here we store the ids of the nodes that have been visited, i.e., we have computed their heuristic
	visited = set()

	# The queue that contains the nodes that we have to visit in the future
	queue = [end_node]

	while queue:
		node_name = queue.pop(0)
		# Check if this node exist in the visited nodes. If not compute its heuristic
		if node_name not in visited:
			# Add node in the vistited nodes
			visited.add(node_name)
			# Take the Object of the examined node
			node = graph.nodes[node_name]
			# Compute its heuristic
			node.heuristic = calculate_heuristic(node_name, end_node, start_node, graph, operator_costs)
			print(f"Visiting node {node_name}, Heuristic: {node.heuristic}")

			# Get the children of the node that we examined earlier
			for child_name in node.get_children():
				# Check if the child is already visited
				if child_name not in visited:
					# Flag in order to check that all the parents of the child have been visited
					all_parents_visited = True
					for parent_child_name in graph.nodes[child_name].get_parents():
						# If there is at least one parent that have not been visited yet then you can not compute the heuristic for this kid
						if parent_child_name not in visited:
							all_parents_visited = False
					if all_parents_visited:
						queue.append(child_name)


def get_communication_cost(graph, child, father, com_lat_mult):
	if graph.nodes[child].get_operator() == 'START':
		return 0

	# Get the location of the two operators
	loc_child = graph.nodes[child].get_location()
	loc_father = graph.nodes[father].get_location()

	# In case they have the same location then the communication cost is zero
	if loc_father == loc_child:
		return 0

	# In case of different locations, exploit the communication graph in order to find the shortest path with its cost
	cost, _ = get_path_cost(COM_GRAPH, loc_child, loc_father)
	return cost * com_lat_mult


# Function that constructs the new plan and adds them to the queue
# so_far_nodes: are the nodes that already exist in the plan that we are studying to extend
# new_node: is the node that we are planning to add
# original_graph: is the topology that was given as input by the user (it is also called guru_graph)
# current_graph: is the graph or plan that we have constructed so far
# possible_locations and possible_frameworks are needed in order to add to the plan all possible combinations of new_node
# END_NODE is needed since the END_NODE does not have many types, i.e., location and framework 
# operator_costs contains the costs of executing any operator to any (valid) device
def construct_new_plan(so_far_nodes, new_node, original_graph, current_graph, possible_locations, possible_frameworks, END_NODE, operator_costs, aggr_func, com_lat_mult):
	# The queue that contains all the plans that we want to study in the future
	tmp_queue_of_plans2 = []

	# All the nodes that we want to insert to the new plan-graph, i.e., $ so_far_nodes \cup new_node $
	all_nodes_for_new_plan = []
	for node in so_far_nodes:
		all_nodes_for_new_plan.append(node)
	all_nodes_for_new_plan.append(new_node)

	# In case that we have the END_NODE we have to avoid the double loop of location-framework since this node does not have any type
	if new_node == END_NODE:
		# Create a new graph object
		tmp_graph = DirectedWeightedGraph()
		# Add each node to the new graph
		for k in range(0, len(all_nodes_for_new_plan)):
			# If this node existed already, i.e., it's not the new_node then take its configurations from the current_graph
			if all_nodes_for_new_plan[k] != new_node:
				tmp_graph.add_node(all_nodes_for_new_plan[k], current_graph.nodes[all_nodes_for_new_plan[k]].get_real_cost(), current_graph.nodes[all_nodes_for_new_plan[k]].get_heuristic_cost(), current_graph.nodes[all_nodes_for_new_plan[k]].get_operator(), current_graph.nodes[all_nodes_for_new_plan[k]].get_location(), current_graph.nodes[all_nodes_for_new_plan[k]].get_framework())
			# If it's the END_NODE then just added to the new graph
			else:
				tmp_graph.add_node(new_node, 0, 0, 'END', 'END', 'END')
		# Once we have inserted all the nodes to the new graph, we need to insert the appropriate edges
		# So, for each node in the new graph 
		for k in range(0, len(all_nodes_for_new_plan)):
			# Get its neighbors (parents) that exist in the ORIGINAL graph
			for neighbor in original_graph.get_parents(all_nodes_for_new_plan[k]):
				# Keep only the parents that exist in the current_graph
				if neighbor in all_nodes_for_new_plan:
					tmp_graph.add_edge(neighbor, all_nodes_for_new_plan[k], 1)

		# Compute the real cost for the just inserted node
		if aggr_func == 'sum':
			################################################
			# I. SUM scenario
			################################################
			sum_of_children_real_costs = 0
			for child in tmp_graph.get_children(new_node):
				sum_of_children_real_costs += tmp_graph.nodes[child].get_real_cost()
			# Get the already stored real cost of this node (this is usefull for the incremental version since there is the case where a node may have data due to the "forward" process)
			tmp = tmp_graph.nodes[new_node].get_real_cost()
			# Set the real cost of the node
			tmp_graph.nodes[new_node].set_real_cost(tmp + sum_of_children_real_costs)
		else:
			###############################################
			# II. MAX scenario
			###############################################
			max_tmp = -1
			for child in tmp_graph.get_children(new_node):
				tmp_val = tmp_graph.nodes[child].get_real_cost()
				if tmp_val >= max_tmp:
					max_tmp = tmp_val
			# Get the already stored real cost of this node (this is usefull for the incremental version since there is the case where a node may have data due to the "forward" process)
			tmp = tmp_graph.nodes[new_node].get_real_cost()
			# Set the real cost of the node
			tmp_graph.nodes[new_node].set_real_cost(tmp + max_tmp)

		# Add the new graph-plan to the queue
		tmp_queue_of_plans2.append(tmp_graph)
	
	# If this is not the END_NODE
	else:
		# We have to crete P new plans where P = N * M, where N is the number of different valid locations, and M is the number of different valid frameworks
		for i in range(0, len(possible_locations)):
			for j in range(0, len(possible_frameworks)):
				# Create a new graph object
				tmp_graph = DirectedWeightedGraph()
				# Add each node to the new graph
				for k in range(0, len(all_nodes_for_new_plan)):
					# If this node existed already, i.e., it's not the new_node then take its configurations from the current_graph
					if all_nodes_for_new_plan[k] != new_node:
						tmp_graph.add_node(all_nodes_for_new_plan[k], current_graph.nodes[all_nodes_for_new_plan[k]].get_real_cost(), current_graph.nodes[all_nodes_for_new_plan[k]].get_heuristic_cost(), current_graph.nodes[all_nodes_for_new_plan[k]].get_operator(), current_graph.nodes[all_nodes_for_new_plan[k]].get_location(), current_graph.nodes[all_nodes_for_new_plan[k]].get_framework())
					# If this is the new_node then add the configuration based on the i and j values (type of location and framework accordingly) in the double loop
					else:
						tmp_graph.add_node(new_node, get_real_cost_of_conf(original_graph.nodes[new_node].get_operator(), i, j, operator_costs), original_graph.nodes[new_node].get_heuristic_cost(), original_graph.nodes[new_node].get_operator(), possible_locations[i], possible_frameworks[j])
				# Once we have inserted all the nodes to the new graph, we need to insert the appropriate edges
				for k in range(0, len(all_nodes_for_new_plan)):
					# Get its neighbors (parents) that exist in the ORIGINAL graph
					for neighbor in original_graph.get_parents(all_nodes_for_new_plan[k]):
						# Keep only the parents that exist in the current_graph
						if neighbor in all_nodes_for_new_plan:
							tmp_graph.add_edge(neighbor, all_nodes_for_new_plan[k], 1)
				
				# Compute the real cost for the just inserted node
				if aggr_func == 'sum':
					################################################
					# I. SUM scenario
					################################################
					sum_of_children_real_costs = 0
					sum_of_children_comunication_costs = 0
					for child in tmp_graph.get_children(new_node):
						sum_of_children_real_costs += tmp_graph.nodes[child].get_real_cost()
						sum_of_children_comunication_costs += get_communication_cost(tmp_graph, child, new_node, com_lat_mult)
					# Get the already stored real cost of this node
					tmp = tmp_graph.nodes[new_node].get_real_cost()
					# Set the real cost of the node
					tmp_graph.nodes[new_node].set_real_cost(tmp + sum_of_children_real_costs + sum_of_children_comunication_costs)
				else:
					###############################################
					# II. MAX scenario
					###############################################
					max_tmp = -1
					for child in tmp_graph.get_children(new_node):
						tmp_val = tmp_graph.nodes[child].get_real_cost() + get_communication_cost(tmp_graph, child, new_node, com_lat_mult)
						if tmp_val >= max_tmp:
							max_tmp = tmp_val
					# Get the already stored real cost of this node
					tmp = tmp_graph.nodes[new_node].get_real_cost()
					# Set the real cost of the node
					tmp_graph.nodes[new_node].set_real_cost(tmp + max_tmp)

				# Add the new graph-plan to the queue
				tmp_queue_of_plans2.append(tmp_graph)
	return tmp_queue_of_plans2

# Given a plan compute its estimated cost
def compute_estimate_cost_of_plan(current_graph):
	min_heuristic_of_leaves = 1000000000000000
	max_real_of_leaves = -1
	#sum_real_of_leaves = 0
	
	# Iterate over the leaves of the current plan
	for i in range(0, len(current_graph.get_root_nodes())):
		# Sum of the real costs of the leaves
		#sum_real_of_leaves += current_graph.nodes[current_graph.get_root_nodes()[i]].get_real_cost()

		# Find the maximum real cost of the leaves of the plan
		if max_real_of_leaves <= current_graph.nodes[current_graph.get_root_nodes()[i]].get_real_cost():
			max_real_of_leaves = current_graph.nodes[current_graph.get_root_nodes()[i]].get_real_cost()

		# Find the minimum heuristic of the leaves of the plan
		if min_heuristic_of_leaves >= current_graph.nodes[current_graph.get_root_nodes()[i]].get_heuristic_cost():
			min_heuristic_of_leaves = current_graph.nodes[current_graph.get_root_nodes()[i]].get_heuristic_cost()
	
	# Compute the estimated cost of the given plan
	estimated_cost_of_plan = max_real_of_leaves + min_heuristic_of_leaves

	return estimated_cost_of_plan

# Given a node and a plan, return all of its fathers, grandfather, grand-grand fathers etc.
def get_all_fathers(plan, node):
	fathers = []
	parents = plan.nodes[node].get_parents()
	
	for parent in parents:
		fathers.append(parent)
		fathers.extend(get_all_fathers(plan, parent))
	
	return fathers


def update_priority_queue(old_queue, nodes_to_change, guru, END_NODE, START_NODE, operator_costs, old_min, new_min):
	# Lists that will be used in order to find the affected-by-changes nodes, their parents, grand-parents etc.
	all_changed_nodes = []
	unique_changed_nodes = []
	
	# Add the nodes that have changed (directly)
	all_changed_nodes.extend(nodes_to_change)

	guru_upd = copy.deepcopy(guru)
	heuristic_flag = False
	# Check if the heuristic was affected
	if new_min != old_min:
		heuristic_flag = True
		for node in guru_upd.get_all_nodes():
			guru_upd.nodes[node].set_real_cost(0)

		print()
		print("============= Old Heuristics =============")
		print(guru)

		# Recompute the heuristics
		bfs(guru_upd, END_NODE, START_NODE, operator_costs)
		
		print()
		print("============= New Heuristics =============")
		print(guru_upd)
		

		# Add to the changed ones all nodes that have different heuristic compared to their heuristic before the changes
		for node in guru_upd.get_all_nodes():
			if guru_upd.nodes[node].get_heuristic_cost() != guru.nodes[node].get_heuristic_cost():
				all_changed_nodes.append(node)

	# Add all the fathers, grand-fathers etc. of the (affected by the changes) nodes 
	for node in nodes_to_change:
		tmp_lst = get_all_fathers(guru, node)
		for f in tmp_lst:
			all_changed_nodes.append(f)

	# Remove duplicates
	all_changed_nodes = list(set(all_changed_nodes))
	#print(all_changed_nodes)

	unique_changed_nodes.extend(all_changed_nodes)
	for node in all_changed_nodes:
		tmp_lst = get_all_fathers(guru, node)
		for f in tmp_lst:
			unique_changed_nodes.append(f)

	# Remove duplicates. 
	# The resulted list contains all the nodes that have to be removed from the sub-plans in queue
	unique_changed_nodes = list(set(unique_changed_nodes))
	print(f"\nThe affected (by the changes) nodes are: {unique_changed_nodes}")

	# The updated priority queue wrt the changes of the dynamic environment
	new_queue = []
	heapq.heapify(new_queue)
	print("\nUpdate the priority queue...")
	for tmp_var in old_queue:
		plan = tmp_var[2]
		intersected_nodes = list(set(unique_changed_nodes).intersection(plan.get_all_nodes()))
		
		# Delete the affected nodes from each subplan along with their edges
		for inter_node in intersected_nodes:
			plan.delete_node(inter_node)

		# Update the heuristic scores of each node
		for node in plan.get_all_nodes():
			if plan.nodes[node].get_heuristic_cost() != guru_upd.nodes[node].get_heuristic_cost():
				plan.nodes[node].set_heuristic_cost(guru_upd.nodes[node].get_heuristic_cost())

		# Set the cost of the plan based on the updated costs
		tmp_cost = compute_estimate_cost_of_plan(plan)
		plan.set_estimated_cost_of_graph(tmp_cost)

		heapq.heappush(new_queue, (plan.estimated_cost, tmp_var[1], plan))

	print(len(old_queue))
	print(len(new_queue))
	# Removing duplicates
	print("\nRemoving the duplicates that were created in the (updated) priority queue...")
	lst_g = map(lambda x: x[2], new_queue)
	set_g = set(lst_g)
	result_new_queue = []
	heapq.heapify(result_new_queue)
	for g in set_g:
		heapq.heappush(result_new_queue, (g.get_estimated_cost_of_graph(), g.get_counterQ(), g))
	print(len(result_new_queue))
	return result_new_queue, guru_upd

def get_k_smallest_costs(objects, k):
    # Sort the list of objects by their estimated_cost attribute
    sorted_objects = sorted(objects, key=lambda obj: obj.estimated_cost)
    # Return the first k objects
    return sorted_objects[:k]

def update_dict_value(dictionary, key, mult_fac):
    """
    Updates the value of a specified key in the dictionary by multiplying 
    each element in the list by a given factor.

    Args:
        dictionary (dict): The dictionary to update.
        key (str): The key to search for in the dictionary.
        mult_fac (float): The multiplication factor.

    Returns:
        dict: The updated dictionary.
    """
    if key in dictionary:
        # Check if the value is a list of floats
        if all(isinstance(item, float) for item in dictionary[key]):
            dictionary[key] = [item * mult_fac for item in dictionary[key]]
        else:
            print(f"Error: The value for key '{key}' is not a list of floats.")
    else:
        print(f"Error: Key '{key}' not found in the dictionary.")

    return dictionary

#########################################
# Driver code
#########################################

# This variable is the #of possible locations (devices in the network) that we want in this simulation
# It is used in order to automatically read the appropriate files without changing anythng else in the code
my_conf_number = 31
mult_factor = 100

# This factor simulates the constant factor for the communication latencies (e.g., 0.1 is the NES-like topology)
com_lat_mult = 1

my_kappa = 1

# The aggregation function that will be empoyed ('sum' or 'max')
aggr_func = 'sum'

# Encoding of the START and END nodes
START_NODE = 1000
END_NODE = 0

# Check if the script is called with the correct argument
if len(sys.argv) > 1:
    # Extract the argument (the filename)
    arg = sys.argv[1]
    
    # Extract the part of the filename before the underscore (the name of the workflow)
    word = arg.split('_')[0]
    
    # Save the word to a variable
    print("The selected workflow is:", word)
else:
    print("No argument provided!")

# Extract the possible locations, i.e., sites for this simulation from the appropriate JSON file
json_file_path = "../network_sample/" + str(word) + "_dataflow/" + str(my_conf_number) + "_1/network_" + str(my_conf_number) + "_1.json"  # Provide the path to your JSON file
site_names = extract_site_names_from_json(json_file_path)

# Extract the information regarding communications latencies for this simulation from the appropriate JSON file
json_file_path = "../network_sample/" + str(word) + "_dataflow/" + str(my_conf_number) + "_1/links_" + str(my_conf_number) + "_1.json"  # Provide the path to your JSON file
from_list, to_list, latency_list = read_json_file(json_file_path)

# Create the communication graph
global COM_GRAPH
COM_GRAPH = create_com_graph(from_list, to_list, latency_list)

# Initialize our original graph (or workflow)
graph = DirectedWeightedGraph()
graph.set_estimated_cost_of_graph(10000)

# Argument parser setup
parser = argparse.ArgumentParser(description="Process a workflow JSON file to create a graph.")
parser.add_argument('json_file', type=str, help="Path to the JSON file containing workflow data")
args = parser.parse_args()

# Read JSON file
with open(args.json_file, 'r') as file:
    data = json.load(file)

# Add nodes
for operator in data['operators']:
    graph.add_node(
        operator['id'], 
        0, 0, 
        operator['name'], 
        'dummy0', 
        'dummy0'
    )

# Add edges
for edge in data['edges']:
    graph.add_edge(edge['from'], edge['to'], 1)

##########################################
##########################################

# Initialize the dictionary
codes_dictionary = {}

# Iterate through the operators in the JSON file
for operator in data['operators']:
    operator_name = operator['name']
    is_fixed = operator['fixed']

    # Build the dictionary entry for each operator
    codes_dictionary[operator_name] = {
        'possible_locations_lst': [site_names[0]] if is_fixed else site_names,
        'possible_frameworks_lst': ["BDP_1"]
    }

# Add 'START' and 'END' nodes
codes_dictionary['START'] = {
    'possible_locations_lst': ['START'],
    'possible_frameworks_lst': ['START']
}
codes_dictionary['END'] = {
    'possible_locations_lst': ['END'],
    'possible_frameworks_lst': ['END']
}

# Read Excel file
excel_file = "../" + str(word) + "_" + str(my_conf_number) + "_dataflow.xlsx"  # Replace with your Excel file path
df = pd.read_excel(excel_file)

# Initialize dictionary to store costs
operator_costs = {}

# Iterate over each row in the Excel file
for _, row in df.iterrows():
    operator_name = row.iloc[0]  # First column contains operator name
    costs = row.iloc[1:].tolist()  # Extract costs from remaining columns

    # Filter valid costs (exclude -1) and multiply them by mult_factor
    valid_costs = [cost * mult_factor for cost in costs if cost != -1]

    # Check if the operator is fixed
    operator_info = next((op for op in data['operators'] if op['name'] == operator_name), None)
    if operator_info and operator_info['fixed']:
        # If fixed, store only the single valid cost as a list
        operator_costs[operator_name] = [valid_costs[0]] if valid_costs else []
    else:
        # Otherwise, store all valid costs
        operator_costs[operator_name] = valid_costs

print(operator_costs)

# Get the leaves of graph
leaf_nodes_lst = graph.get_leaf_nodes()

# Get the roots of graph
root_nodes_lst = graph.get_root_nodes()

# Add in an automatic way the END and START nodes to the graph
graph.add_node(END_NODE, 0, 0, 'END', 'END', 'END')
graph.add_node(START_NODE, 0, 0, 'START', 'START', 'START')

# Add the edges to the END and START nodes in an automatic way
for root in root_nodes_lst:
	graph.add_edge(END_NODE, root, 1)
for leaf in leaf_nodes_lst:
	graph.add_edge(leaf, START_NODE, 1)

# Print the details of the "original" workflow
print("My Graph Structure (or Workflow) before defining the heuristics...:")
print(graph)

# Call the bfs() function for calculating the heuristics 
bfs(graph, END_NODE, START_NODE, operator_costs)

print("\n=================================")
print("Heuristics were computed succesfully...")
print("=================================")
print("My Graph Structure (or Workflow):")
print(graph)

# Define the queue that we will store the plans based on the A*-alike algorithm (currently we treat this variable as a GLOBAL one)
queue_of_plans = []

# Indication that currently we are at the very first iteration
counter = -1

# Add the first plan to the queue
# The first plan is a graph that contains only the START node
print("\n=================================")
print("start graph...")
print("=================================")
graph_start = DirectedWeightedGraph()
graph_start.add_node(START_NODE, 0, 0, 'START', 'START', 'START')
graph_start.set_estimated_cost_of_graph(100000000000000)
graph_start.nodes[START_NODE].set_heuristic_cost(0)
graph_start.nodes[START_NODE].set_real_cost(0)

counter_heapQ = 1
heapq.heappush(queue_of_plans, (100000000000000, counter_heapQ, graph_start))
graph_start.set_counterQ(counter_heapQ)

# The ground truth, based on this we make the configurations, get the parents-kids etc. in our A*-alike search of the optimal solution
graph_guru = graph

while True:
	print("\n=================================")
	print("Searching the Optimal Plan...")
	print("=================================")
	# Set counter to zero in order to compute the iterations that we will need in order to find the optimal plan
	counter = 0
	while len(queue_of_plans) > 0:
		counter += 1

		# This flag indicates whether an extension was made. If its True, it means that an extension was made and we have to break and continue with the next iteration (remember Deligiannakis meeting)
		added_flag = False
		
		# Here we will store the newly constucted sub-plan, i.e., the sub-plans that were generated after the extension
		tmp_queue_of_plans = []

		# Dequeue the next plan
		x1, x2, plan_tmp = heapq.heappop(queue_of_plans)
		print(counter)
		print(f"The estimated cost of the examined plan is ----> {plan_tmp.get_estimated_cost_of_graph()}")

		# Check if there is no node left to add, i.e., the root of the examined plan is the END_NODE
		if END_NODE in plan_tmp.get_root_nodes():
			print("Success: Found best plan...")
			break

		# Get all nodes of the examined plan
		nodes_plan_tmp = plan_tmp.get_all_nodes()
		
		# For each node in the current plan find its parent(s)
		parent_nodes_plan_tmp = []
		for nodes in nodes_plan_tmp:
			parent_nodes_plan_tmp.append(graph_guru.get_parents(nodes))

		# For each node-parent of the existing nodes in our plan find which are valid, i.e., in order to add a new node all of its children have to be included in the current plan
		# else its no-valid and we can not add the parent-node to the current plan
		# Moreover we need to check the OTHER_CONDITION as well...
		for i in range(0, len(parent_nodes_plan_tmp)):
			if added_flag:
				break
			for j in range(0, len(parent_nodes_plan_tmp[i])):
				tmp = graph_guru.get_children(parent_nodes_plan_tmp[i][j])
				flag = True
				for k in range(0, len(tmp)):
					if tmp[k] not in nodes_plan_tmp:
						flag = False
				
				# OTHER_CONDITION: If the parent of any node already exists, e.g., the parents of START node already exists after the first iterations(s)
				if parent_nodes_plan_tmp[i][j] in nodes_plan_tmp:
					flag = False
				
				# If all of parent-node children have been already added in the current plan then add the parent and the OTHER_CONDITION is met then add the new plan
				if flag:
					# Find the valid locations and frameworks based on the CODE, i.e., type of operator, that the node-to-be-added has
					possible_locations = codes_dictionary[graph_guru.nodes[parent_nodes_plan_tmp[i][j]].get_operator()]['possible_locations_lst']
					possible_frameworks = codes_dictionary[graph_guru.nodes[parent_nodes_plan_tmp[i][j]].get_operator()]['possible_frameworks_lst']
					tmp_queue_of_plans = construct_new_plan(nodes_plan_tmp, parent_nodes_plan_tmp[i][j], graph_guru, plan_tmp, possible_locations, possible_frameworks, END_NODE, operator_costs, aggr_func, com_lat_mult)

					# Compute the estimated costs of the newly inserted plans
					# If the-just-inserted-node is the END node then we added only one plan which is in the last position of the queue (since it hasn't been sorted yet)
					if parent_nodes_plan_tmp[i][j] == END_NODE:
						tmp_cost = compute_estimate_cost_of_plan(tmp_queue_of_plans[-1])
						tmp_queue_of_plans[-1].set_estimated_cost_of_graph(tmp_cost)
					else:
						# In case that the new node isn't the END node then there are possible_locations * possible_frameworks new plans that
						# were added in the queue via the construct_new_plan() function
						# So, for each of these new plans compute their estimated cost
						for g in tmp_queue_of_plans:
							tmp_cost = compute_estimate_cost_of_plan(g)
							g.set_estimated_cost_of_graph(tmp_cost)
					# Update the flag in order to break and continue with the sorting process
					added_flag = True
					break

		if len(tmp_queue_of_plans) > 1:
			smallest_cost_items_sorted = get_k_smallest_costs(tmp_queue_of_plans, my_kappa)
			del tmp_queue_of_plans
			tmp_queue_of_plans = smallest_cost_items_sorted

		for plan_tmp2 in tmp_queue_of_plans:
			counter_heapQ += 1
			heapq.heappush(queue_of_plans, (plan_tmp2.estimated_cost, counter_heapQ, plan_tmp2))
			plan_tmp2.set_counterQ(counter_heapQ)

		##############################################################
		# If we want to have a priority queue with fixed size. NO GUARANTEE FOR OPTIMALITY
		# queue_of_plans = queue_of_plans[:5000]
		# print(f"The size of queue currently is -----> {len(queue_of_plans)} ...")
		##############################################################

	# Print the number of iterations that we performed in order to find the optimal solution
	print("\n=================================================================================================")
	print(f"We performed {counter} iterations in order to discover the optimal solution...")
	print(f"The size of queue currently is -----> {len(queue_of_plans)} ...")

	# Store the just computed optimal plan for thsi simulation
	opt_plan = plan_tmp
	print("OPTIMAL PLAN:")
	print(opt_plan)

	# The exit code for our simulation
	zero_code = int(input('Press 0 for quiting: '))
	if zero_code == 0:
		break

	##############################################################
	# Incremental A*-alike part (d-DAG*)
	##############################################################

	# Insert again the optimal plan in order to guarantee that all possible confs are available for the incremental version
	heapq.heappush(queue_of_plans, (opt_plan.estimated_cost, x2, opt_plan))

	##############################################################
	# User defined change
	##############################################################

	nodes_to_change = set()
	print(operator_costs)
	affected_operator = input("Enter the affected operator:")
	multiplier = float(input("Enter the multiplication factor: "))

	for node in graph.get_all_nodes():
		if graph.nodes[node].get_operator() == affected_operator:
			nodes_to_change.add(node)	
	print(nodes_to_change)

	old_min = float(min(operator_costs[affected_operator]))
	operator_costs = update_dict_value(operator_costs, affected_operator, multiplier)
	print("Updated dictionary:", operator_costs)
	new_min = float(min(operator_costs[affected_operator]))

	new_priority_queue, graph_guru = update_priority_queue(queue_of_plans, nodes_to_change, graph_guru, END_NODE, START_NODE, operator_costs, old_min, new_min)

	queue_of_plans.clear()
	
	for obj in new_priority_queue:
		heapq.heappush(queue_of_plans, (obj[0], obj[1], obj[2]))

print("Terminating A*-alike simulation...")